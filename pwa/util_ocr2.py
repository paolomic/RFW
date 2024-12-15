import pyautogui
import openpyxl
#from pywinauto.uia_defines import IUIA, UIAControl

def capture_grid_as_image(grid_control):
    """
    Cattura un'immagine della griglia specificata
    """
    # Ottieni il rettangolo delimitante del controllo griglia
    rect = grid_control.rectangle()
    
    # Cattura uno screenshot dell'area della griglia
    return pyautogui.screenshot(region=(rect.left, rect.top, rect.width(), rect.height()))

def save_grid_to_excel(grid_data, file_path):
    """
    Salva i dati della griglia in un file Excel
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    for row, row_data in enumerate(grid_data, start=1):
        for col, cell_value in enumerate(row_data, start=1):
            worksheet.cell(row=row, column=col, value=cell_value)

    workbook.save(file_path)

def extract_grid_data_from_control(grid_control):
    """
    Estrae i dati della griglia dal controllo griglia specificato
    """
    # Cattura l'immagine dell'area della griglia
    grid_image = capture_grid_as_image(grid_control)
    
    # Salva l'immagine in un file temporaneo
    grid_image.save('grid_screenshot.png')
    
    # Chiedi all'utente di copiare manualmente il contenuto della griglia
    #print("Copia manualmente il contenuto della griglia in un file Excel.")
    
    # Aspetta che l'utente confermi di aver copiato i dati
    #input("Premi Invio dopo aver copiato i dati.")
    
    # Leggi i dati dal file Excel creato dall'utente
    grid_data = read_excel_data('grid_data.xlsx')
    
    return grid_data

def read_excel_data(file_path):
    """
    Legge i dati dalla griglia dal file Excel
    """
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    
    grid_data = []
    for row in worksheet.iter_rows(values_only=True):
        grid_data.append(row)
    
    return grid_data

